import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import openpyxl
from datetime import datetime
import requests
import io

st.set_page_config(
    page_title="Fuerza de Trabajo Notificada - C.C.C. Dos Bocas",
    layout="wide"
)
# ── ESTADO ─────────────────────────────────────────────
if "ver_mec" not in st.session_state:
    st.session_state.ver_mec = False
if "ver_ele" not in st.session_state:
    st.session_state.ver_ele = False
if "ver_instyctrl" not in st.session_state:
    st.session_state.ver_instyctrl = False
if "ver_civ" not in st.session_state:
    st.session_state.ver_civ = False

st.markdown("""
<style>
    .main-title { font-size: 26px; font-weight: bold; color: #1a3a5c; margin-bottom: 4px; }
    .sub-title  { font-size: 14px; color: #555; margin-bottom: 20px; }
    .kpi-box    { background: #f0f4f8; border-radius: 10px; padding: 16px 20px; text-align: center; }
    .kpi-label  { font-size: 13px; color: #666; margin-bottom: 4px; }
    .kpi-value  { font-size: 28px; font-weight: bold; }
    .section-header { font-size: 18px; font-weight: bold; color: #1a3a5c; margin: 24px 0 10px 0; }
</style>
""", unsafe_allow_html=True)



# ID del archivo en Google Drive
FILE_ID = "1_IHIe0xIj3kmKgNG43a3k4cmACm5o7ob"
DRIVE_URL = f"https://drive.google.com/uc?export=download&id={FILE_ID}"

def leer_hoja_trabajadores(wb, nombre_hoja):
    ws = wb[nombre_hoja]
    rows = list(ws.iter_rows(min_row=1, max_row=100, values_only=True))
    # 🔍 1. Detectar encabezado (donde empiezan las fechas)
    header_idx = None
    for i, row in enumerate(rows):
        if row and any(isinstance(c, datetime) for c in row):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(f"No se encontró encabezado en hoja {nombre_hoja}")

    header_row = rows[header_idx]

    # 📅 2. Detectar columnas de fechas automáticamente
    date_headers = []
    start_col = None

    for i, val in enumerate(header_row):
        if isinstance(val, datetime):
            start_col = i
            break

    for val in header_row[start_col:]:
        if isinstance(val, datetime):
            date_headers.append(val.strftime('%d/%m'))
        elif val:
            date_headers.append(str(val))
        else:
            break  # 🔥 se detiene cuando ya no hay fechas

   # 👤 3. Detectar columnas automáticamente
    headers_text = [
    str(c).strip().upper() if c else ''
    for c in header_row
    ]

    col_nombre = headers_text.index('NOMBRE')
    col_rpe = headers_text.index('RPE')
    col_categ = headers_text.index('CATEGORIA')

    # 👤 4. Leer personas
    people = []

    for r in rows[header_idx + 1:]:
        if not r or len(r) < start_col:
         continue

        nombre = r[col_nombre] if len(r) > col_nombre else ''
        rpe    = r[col_rpe] if len(r) > col_rpe else ''
        categ  = r[col_categ] if len(r) > col_categ else ''

    # filtros
        if not nombre or str(nombre).startswith('   SUPLENTES') or nombre == 'VACANTE':
         continue

    # 🧮 horas dinámicas
    horas = [
        v if isinstance(v, (int, float)) else 0
        for v in r[start_col:start_col + len(date_headers)]
    ]

    people.append({
        'Nombre': nombre,
        'RPE': rpe or '',
        'Categoría': categ or '',
        'Total_hrs': sum(horas),
        **{
            date_headers[i]: horas[i]
            for i in range(len(date_headers))
        }
    })

    df = pd.DataFrame(people)

    return df, date_headers
    def color_total(val):
        try:
            v = float(val)
            max_v = df_civ_fil['Total_hrs'].max() or 1
            intensity = int(200 - (v / max_v) * 150)
            return f'background-color: rgb({intensity}, {intensity+30}, 255); color: {"white" if intensity < 100 else "black"}'
        except:
            return ''

@st.cache_data(ttl=300)  # refresca cada 5 minutos
def load_data():
    # Descargar el Excel desde Google Drive
    session = requests.Session()
    response = session.get(DRIVE_URL, stream=True)

    # Google Drive a veces muestra una página de confirmación para archivos grandes
    for key, value in response.cookies.items():
        if key.startswith("download_warning"):
            response = session.get(DRIVE_URL, params={"confirm": value}, stream=True)
            break

    excel_bytes = io.BytesIO(response.content)
    wb = openpyxl.load_workbook(excel_bytes, data_only=True)

    # ── RESUMEN_OTS ──────────────────────────────────────────────────────────
    ws_res = wb['RESUMEN_OTS']
    rows_res = list(ws_res.iter_rows(values_only=True))

    anio = rows_res[2][2]
    mes  = rows_res[3][2]
    fecha_act = rows_res[3][5]
    if isinstance(fecha_act, datetime):
        fecha_act = fecha_act.strftime('%d/%m/%Y')

    dept_rows = rows_res[7:11]
    df_resumen = pd.DataFrame(
        [(r[1], r[2], r[3], r[4], r[5]) for r in dept_rows if r[1]],
        columns=['Departamento', 'Cumplimiento', 'Disponible', 'Planificado', 'Notificado']
    )
    df_resumen['Cumplimiento_pct'] = (df_resumen['Cumplimiento'] * 100).round(2)

    total_row = rows_res[12]
    totales = {
        'cumplimiento': total_row[2],
        'disponible':   total_row[3],
        'planificado':  total_row[4],
        'notificado':   total_row[5],
    }

    # ── MEC ──────────────────────────────────────────────────────────────────
    df_mec, date_headers = leer_hoja_trabajadores(wb, 'MEC')

    # ── ELECTRICO ──────────────────────────────────────────────────────────────────
    df_ele, date_headers_ele = leer_hoja_trabajadores(wb, 'ELE')
    # INSTYCTRL_______________________________________________________________________
    df_instyctrl, date_headers_instyctrl = leer_hoja_trabajadores(wb, 'IYC')
    # CIV_______________________________________________________________________
    df_civ, date_headers_civ = leer_hoja_trabajadores(wb, 'CIV')
    

    return (
        anio,
        mes,
        fecha_act,
        df_resumen,
        totales,
        df_mec,
        date_headers,
        df_ele,
        date_headers_ele,
        df_instyctrl,
        date_headers_instyctrl,
        df_civ,
        date_headers_civ
    )
# ── CARGA DE DATOS ────────────────────────────────────────────────────────────
try:
    with st.spinner("Cargando datos desde Google Drive..."):
        anio, mes, fecha_act, df_resumen, totales, df_mec, date_headers, df_ele, data_headers_ele, df_instyctrl, data_headers_instyctrl, df_civ, data_headers_civ = load_data()
except Exception as e:
    st.error(f" No se pudo cargar el archivo desde Google Drive.\n\nVerifica que el archivo sea público.\n\nError: {e}")
    st.stop()

# ── ENCABEZADO ────────────────────────────────────────────────────────────────
st.markdown('<div class="main-title"> Resumen de Cumplimiento de Fuerza de Trabajo Notificada</div>', unsafe_allow_html=True)
st.markdown(f'<div class="sub-title">C.C.C. Dos Bocas &nbsp;|&nbsp; Año: <b>{anio}</b> &nbsp;|&nbsp; Mes: <b>{mes}</b> &nbsp;|&nbsp; Actualizado al: <b>{fecha_act}</b></div>', unsafe_allow_html=True)

# ── KPIs ─────────────────────────────────────────────────────────────────────
k1, k2, k3, k4 = st.columns(4)
cumpl_color = "#27ae60" if totales['cumplimiento'] >= 0.85 else "#e67e22" if totales['cumplimiento'] >= 0.70 else "#e74c3c"
with k1:
    st.markdown(f"""<div class="kpi-box">
        <div class="kpi-label">Cumplimiento Total</div>
        <div class="kpi-value" style="color:{cumpl_color}">{totales['cumplimiento']*100:.1f}%</div>
    </div>""", unsafe_allow_html=True)
with k2:
    st.markdown(f"""<div class="kpi-box">
        <div class="kpi-label">Horas Disponibles</div>
        <div class="kpi-value" style="color:#2980b9">{totales['disponible']:,.1f}</div>
    </div>""", unsafe_allow_html=True)
with k3:
    st.markdown(f"""<div class="kpi-box">
        <div class="kpi-label">Horas Planificadas</div>
        <div class="kpi-value" style="color:#8e44ad">{totales['planificado']:,.1f}</div>
    </div>""", unsafe_allow_html=True)
with k4:
    st.markdown(f"""<div class="kpi-box">
        <div class="kpi-label">Horas Notificadas</div>
        <div class="kpi-value" style="color:#16a085">{totales['notificado']:,.1f}</div>
    </div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ── GRÁFICA ───────────────────────────────────────────────────────────────────
st.markdown('<div class="section-header"> Cumplimiento por Departamento</div>', unsafe_allow_html=True)

fig = go.Figure()
fig.add_trace(go.Bar(
    name='Disponible', x=df_resumen['Departamento'], y=df_resumen['Disponible'],
    marker_color='#29b978', text=df_resumen['Disponible'].apply(lambda v: f"{v:,.0f}"), textposition='outside',
))
fig.add_trace(go.Bar(
    name='Planificado', x=df_resumen['Departamento'], y=df_resumen['Planificado'],
    marker_color='#635f65', text=df_resumen['Planificado'].apply(lambda v: f"{v:,.0f}"), textposition='outside',
))
fig.add_trace(go.Bar(
    name='Notificado', x=df_resumen['Departamento'], y=df_resumen['Notificado'],
    marker_color='#c3c7c5', text=df_resumen['Notificado'].apply(lambda v: f"{v:,.0f}"), textposition='outside',
))
fig.add_trace(go.Scatter(
    name='Cumplimiento (%)', x=df_resumen['Departamento'], y=df_resumen['Cumplimiento_pct'],
    mode='lines+markers+text', yaxis='y2',
    line=dict(color='#e74c3c', width=3), marker=dict(size=10, color='#e74c3c'),
    text=df_resumen['Cumplimiento_pct'].apply(lambda v: f"{v:.1f}%"),
    textposition='top center', textfont=dict(color='#e74c3c', size=12),
))
fig.update_layout(
    barmode='group',
    yaxis=dict(title='Horas', gridcolor='#eee'),
    yaxis2=dict(title='Cumplimiento (%)', overlaying='y', side='right',
                range=[0, 130], ticksuffix='%', showgrid=False),
    legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
    plot_bgcolor='white', paper_bgcolor='white',
    height=460, margin=dict(t=40, b=20), font=dict(family='Arial'),
)
st.plotly_chart(fig, use_container_width=True)

# ── TABLA RESUMEN ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-header"> Tabla Resumen – Hoja RESUMEN_OTS</div>', unsafe_allow_html=True)

df_display = df_resumen[['Departamento', 'Disponible', 'Planificado', 'Notificado', 'Cumplimiento_pct']].copy()
df_display.columns = ['Departamento', 'Disponible (Hrs)', 'Planificado (Hrs)', 'Notificado (Hrs)', 'Cumplimiento (%)']
st.dataframe(
    df_display,
    use_container_width=True,
    hide_index=True,
)

# ── TABLA MEC ─────────────────────────────────────────────────────────────────
if st.session_state.ver_mec:

    st.markdown('<div class="section-header"> Detalle de Horas Notificadas – Hoja MEC (Departamento Mecánico)</div>', unsafe_allow_html=True)

    col_busq, col_cat = st.columns([2, 2])
    with col_busq:
        busqueda = st.text_input(" Buscar nombre o RPE", "")
    with col_cat:
        categorias = ['Todas'] + sorted(df_mec['Categoría'].dropna().unique().tolist())
        cat_sel = st.selectbox("Filtrar por categoría", categorias)

    df_mec_fil = df_mec.copy()

    if busqueda:
        df_mec_fil = df_mec_fil[
            df_mec_fil['Nombre'].str.contains(busqueda, case=False, na=False) |
            df_mec_fil['RPE'].astype(str).str.contains(busqueda, case=False, na=False)
        ]

    if cat_sel != 'Todas':
        df_mec_fil = df_mec_fil[df_mec_fil['Categoría'] == cat_sel]

    valid_dates = [d for d in date_headers if d]
    cols_show = ['Nombre', 'RPE', 'Categoría', 'Total_hrs'] + valid_dates
    df_mec_fil = df_mec_fil[[c for c in cols_show if c in df_mec_fil.columns]]


    st.dataframe(df_mec_fil, use_container_width=True)

    st.caption(f"Mostrando {len(df_mec_fil)} de {len(df_mec)} personas")

# ── TABLA ELE ─────────────────────────────────────────────────────────────────
if st.session_state.ver_ele:

    st.markdown('<div class="section-header"> Detalle de Horas Notificadas – Hoja ELE (Departamento Electrico)</div>', unsafe_allow_html=True)

    col_busq, col_cat = st.columns([2, 2])
    with col_busq:
        busqueda = st.text_input(" Buscar nombre o RPE", key="busqueda_ele")
    with col_cat:
        categorias = ['Todas'] + sorted(df_ele['Categoría'].dropna().unique().tolist())
        cat_sel = st.selectbox("Filtrar por categoría", categorias, key="cat_ele")

    df_ele_fil = df_ele.copy()

    if busqueda:
        df_ele_fil = df_ele_fil[
            df_ele_fil['Nombre'].str.contains(busqueda, case=False, na=False) |
            df_ele_fil['RPE'].astype(str).str.contains(busqueda, case=False, na=False)
        ]

    if cat_sel != 'Todas':
        df_ele_fil = df_ele_fil[df_ele_fil['Categoría'] == cat_sel]

    valid_dates = [d for d in data_headers_ele if d]
    cols_show = ['Nombre', 'RPE', 'Categoría', 'Total_hrs'] + valid_dates
    df_ele_fil = df_ele_fil[[c for c in cols_show if c in df_ele_fil.columns]]


    st.dataframe(df_ele_fil, use_container_width=True)

    st.caption(f"Mostrando {len(df_ele_fil)} de {len(df_ele)} personas")

    # ── TABLA INSTYCTRL ─────────────────────────────────────────────────────────────────
if st.session_state.ver_instyctrl:

    st.markdown('<div class="section-header"> Detalle de Horas Notificadas – Hoja INSTYCTRL (Departamento Inst y Ctrl)</div>', unsafe_allow_html=True)

    col_busq, col_cat = st.columns([2, 2])
    with col_busq:
        busqueda = st.text_input(" Buscar nombre o RPE", key="busqueda_instyctrl")
    with col_cat:
        categorias = ['Todas'] + sorted(df_instyctrl['Categoría'].dropna().unique().tolist())
        cat_sel = st.selectbox("Filtrar por categoría", categorias, key="cat_instyctrl")

    df_instyctrl_fil = df_instyctrl.copy()

    if busqueda:
        df_instyctrl_fil = df_instyctrl_fil[
            df_instyctrl_fil['Nombre'].str.contains(busqueda, case=False, na=False) |
            df_instyctrl_fil['RPE'].astype(str).str.contains(busqueda, case=False, na=False)
        ]

    if cat_sel != 'Todas':
        df_instyctrl_fil = df_instyctrl_fil[df_instyctrl_fil['Categoría'] == cat_sel]

    valid_dates = [d for d in data_headers_instyctrl if d]
    cols_show = ['Nombre', 'RPE', 'Categoría', 'Total_hrs'] + valid_dates
    df_instyctrl_fil = df_instyctrl_fil[[c for c in cols_show if c in df_instyctrl_fil.columns]]


    st.dataframe(df_instyctrl_fil, use_container_width=True)

    st.caption(f"Mostrando {len(df_instyctrl_fil)} de {len(df_instyctrl)} personas")

    # ── TABLA CIV ─────────────────────────────────────────────────────────────────
if st.session_state.ver_civ:

    st.markdown('<div class="section-header"> Detalle de Horas Notificadas – Hoja civ (Departamento civ)</div>', unsafe_allow_html=True)

    col_busq, col_cat = st.columns([2, 2])
    with col_busq:
        busqueda = st.text_input(" Buscar nombre o RPE", key="busqueda_civ")
    with col_cat:
        categorias = ['Todas'] + sorted(df_civ['Categoría'].dropna().unique().tolist())
        cat_sel = st.selectbox("Filtrar por categoría", categorias, key="cat_civ")

    df_civ_fil = df_civ.copy()

    if busqueda:
        df_civ_fil = df_civ_fil[
            df_civ_fil['Nombre'].str.contains(busqueda, case=False, na=False) |
            df_civ_fil['RPE'].astype(str).str.contains(busqueda, case=False, na=False)
        ]

    if cat_sel != 'Todas':
        df_civ_fil = df_civ_fil[df_civ_fil['Categoría'] == cat_sel]

    valid_dates = [d for d in data_headers_civ if d]
    cols_show = ['Nombre', 'RPE', 'Categoría', 'Total_hrs'] + valid_dates
    df_civ_fil = df_civ_fil[[c for c in cols_show if c in df_civ_fil.columns]]


    st.dataframe(df_civ_fil, use_container_width=True)

    st.caption(f"Mostrando {len(df_civ_fil)} de {len(df_civ)} personas")




# ── FUNCIONES DE BOTONES ──────────────────────────────────────────
def activar_mec():
    st.session_state.ver_mec= not st.session_state.ver_mec
def activar_ele():
    st.session_state.ver_ele= not st.session_state.ver_ele  
def activar_instyctrl():
    st.session_state.ver_instyctrl= not st.session_state.ver_instyctrl
def activar_civ():
    st.session_state.ver_civ= not st.session_state.ver_civ
# ── SIDEBAR ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("C.C.C. Dos Bocas")
    st.caption("Fuerza de Trabajo Notificada")


    st.selectbox("MES", ["ENERO", "FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE", "NOVIEMBRE","DICIEMBRE"])
    st.button ("MECANICO",on_click=activar_mec)
    st.button ("ELECTRICO",on_click=activar_ele)
    st.button ("INSTYCTRL", on_click=activar_instyctrl)
    st.button ("CIVIL", on_click=activar_civ)
    st.markdown("---")

    if st.button(" Actualizar datos"):
        st.cache_data.clear()
        st.rerun()
    st.markdown("---")
    st.caption(" Datos cargados desde Google Drive")

